from openpyxl import load_workbook

# Load the Excel file
wb = load_workbook('keerthana.xlsx')

# Select the active worksheet
sheet = wb.active

# Read data from cells
data = []

# Iterate over rows and append data to the list
for row in sheet.iter_rows(values_only=True):
    data.append(row)

# Display the data
for row in data:
    print(row)
